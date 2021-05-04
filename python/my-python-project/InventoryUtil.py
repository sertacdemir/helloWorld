import openpyxl

inv_file = openpyxl.load_workbook("doc/inventory.xlsx")
product_list = inv_file["Sheet1"]

num_products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value

    # number of product per supplier
    if supplier_name in num_products_per_supplier:
        current_num = num_products_per_supplier.get(supplier_name)
        num_products_per_supplier[supplier_name] = current_num + 1
    else:
        num_products_per_supplier[supplier_name] = 1

    # Calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_number)] = int(inventory)

    # add total price to each row
    inventory_price = product_list.cell(product_row, 5)
    inventory_price.value = inventory * price

print(num_products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

# save the file with total calculation
product_list.cell(1, 5).value = "Total Price"
product_list.cell(1, 5).font = openpyxl.styles.Font(bold=True)
inv_file.save("doc/inventory_total.xlsx")
